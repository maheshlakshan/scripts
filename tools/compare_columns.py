#!/usr/bin/env python3
"""
Compare columns between two CSV or XLSX files using a column mapping.

Dependencies:
  CSV files  → none (stdlib only)
  XLSX files → openpyxl  (pip install openpyxl)

Usage examples:

  # Discover column names before building a mapping
  python compare_columns.py file1.csv file2.xlsx --map x:x --list-columns

  # Positional comparison (row 0 vs row 0, row 1 vs row 1, ...)
  python compare_columns.py file1.csv file2.xlsx \\
    --map "sku:SKU" "name:product_name"

  # Key-based comparison (match rows by a key column, handles reordering)
  python compare_columns.py file1.csv file2.xlsx \\
    --map "sku:SKU" "name:product_name" "price:unit_price" \\
    --key1 sku --key2 SKU

  # Same key column name in both files
  python compare_columns.py file1.csv file2.csv \\
    --map "price:price" "qty:qty" \\
    --key sku

  # Ignore case and surrounding whitespace when comparing
  python compare_columns.py file1.csv file2.csv \\
    --map "name:Name" --normalize

  # Print only mismatched rows in the terminal
  python compare_columns.py file1.csv file2.xlsx \\
    --map "sku:SKU" "price:Price" \\
    --key1 sku --key2 SKU \\
    --mismatches-only

  # Save full diff report to CSV or XLSX
  python compare_columns.py file1.csv file2.xlsx \\
    --map "sku:SKU" "name:product_name" \\
    --key1 sku --key2 SKU \\
    --output diff_report.csv

  # Filter rows before comparing (reduce comparison range)
  python compare_columns.py order_guide.csv vendor.xlsx \\
    --map "item_no:Item Code*" \\
    --key1 item_no --key2 "Item Code*" \\
    --filter2 "cust_no=40007" \\
    --filter1 "Section~produce"

  # Multiple filters on the same file (all must match — AND logic)
  python compare_columns.py file1.csv file2.csv \\
    --map "sku:sku" \\
    --filter1 "region=West" "status!=inactive"

Filter expression syntax:
  col=value      exact match (case-insensitive with --normalize)
  col!=value     not equal
  col~value      contains  (always case-insensitive)
  col!~value     does not contain
  col^value      starts with (always case-insensitive)
  col$value      ends with   (always case-insensitive)

Column mapping format:  --map "col_in_file1:col_in_file2" ...
"""

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# ANSI colour helpers
# ---------------------------------------------------------------------------

RESET  = "\033[0m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
RED    = "\033[31m"
GREEN  = "\033[32m"
YELLOW = "\033[33m"
CYAN   = "\033[36m"

SUPPORTED = {".csv", ".xlsx", ".xls"}

Row     = Dict[str, str]
Table   = Tuple[List[str], List[Row]]   # (headers, rows)
Report  = List[Dict]


def _c(text, *codes: str) -> str:
    """Wrap text in ANSI codes only when stdout is a real terminal."""
    if sys.stdout.isatty():
        return "".join(codes) + str(text) + RESET
    return str(text)


# ---------------------------------------------------------------------------
# File I/O — no pandas, no numpy
# ---------------------------------------------------------------------------

def _load_csv(path: str) -> Table:
    try:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            rows = [dict(r) for r in reader]
            headers = list(reader.fieldnames or [])
        return headers, rows
    except Exception as exc:
        sys.exit(f"Cannot read {path}: {exc}")


def _load_xlsx(path: str, sheet: Optional[str] = None) -> Table:
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "openpyxl is required for XLSX/XLS files.\n"
            "Install it with:  pip install openpyxl"
        )
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        row_iter = ws.iter_rows(values_only=True)
        raw_headers = next(row_iter, None)
        if raw_headers is None:
            return [], []
        headers = [str(h) if h is not None else "" for h in raw_headers]
        rows: List[Row] = []
        for raw in row_iter:
            rows.append(
                {headers[i]: (str(v) if v is not None else "")
                 for i, v in enumerate(raw) if i < len(headers)}
            )
        wb.close()
        return headers, rows
    except Exception as exc:
        sys.exit(f"Cannot read {path}: {exc}")


def _load(path: str, sheet: Optional[str] = None) -> Table:
    p = Path(path)
    if not p.exists():
        sys.exit(f"File not found: {path}")
    ext = p.suffix.lower()
    if ext not in SUPPORTED:
        sys.exit(
            f"Unsupported extension '{ext}' for {path}. "
            f"Supported: {', '.join(sorted(SUPPORTED))}"
        )
    return _load_csv(path) if ext == ".csv" else _load_xlsx(path, sheet)


# ---------------------------------------------------------------------------
# Mapping + validation helpers
# ---------------------------------------------------------------------------

def _parse_mapping(items: List[str]) -> List[Tuple[str, str]]:
    pairs = []
    for item in items:
        parts = item.split(":", 1)
        if len(parts) != 2 or not parts[0].strip() or not parts[1].strip():
            sys.exit(
                f"Invalid mapping '{item}'. "
                "Expected format: 'file1_column:file2_column'"
            )
        pairs.append((parts[0].strip(), parts[1].strip()))
    return pairs


def _check_cols(headers: List[str], needed: List[str], label: str) -> None:
    missing = [c for c in needed if c not in headers]
    if missing:
        sys.exit(
            f"Column(s) not found in {label}: {', '.join(missing)}\n"
            f"Available: {', '.join(headers)}"
        )


def _norm(v: str) -> str:
    return v.strip().lower()


# ---------------------------------------------------------------------------
# Row filtering
# ---------------------------------------------------------------------------

# Supported operators parsed longest-first to avoid prefix collisions
_OPS = ["!=", "!~", "^", "$", "~", "="]


def _parse_filters(exprs: List[str], headers: List[str], label: str) -> List[Tuple]:
    """
    Parse filter expressions into (col, op, value) tuples.
    Validates that each referenced column exists in headers.
    """
    parsed = []
    for expr in exprs:
        matched_op = None
        for op in _OPS:
            idx = expr.find(op)
            if idx > 0:
                col = expr[:idx].strip()
                val = expr[idx + len(op):]
                matched_op = op
                break
        if matched_op is None:
            sys.exit(
                f"Invalid filter '{expr}'. "
                f"Expected format: col=value, col!=value, col~value, col!~value, col^value, or col$value"
            )
        if col not in headers:
            sys.exit(
                f"Filter column '{col}' not found in {label}.\n"
                f"Available: {', '.join(headers)}"
            )
        parsed.append((col, matched_op, val))
    return parsed


def _apply_filters(rows: List[Row], filters: List[Tuple]) -> List[Row]:
    """Keep only rows that satisfy every filter (AND logic)."""
    if not filters:
        return rows

    def _matches(row: Row, col: str, op: str, val: str) -> bool:
        cell = row.get(col, "")
        cell_l = cell.lower()
        val_l  = val.lower()
        if op == "=":
            return cell_l == val_l
        if op == "!=":
            return cell_l != val_l
        if op == "~":
            return val_l in cell_l
        if op == "!~":
            return val_l not in cell_l
        if op == "^":
            return cell_l.startswith(val_l)
        if op == "$":
            return cell_l.endswith(val_l)
        return False

    return [r for r in rows if all(_matches(r, col, op, val) for col, op, val in filters)]


# ---------------------------------------------------------------------------
# Core comparison
# ---------------------------------------------------------------------------

def compare(
    headers1: List[str],
    rows1: List[Row],
    headers2: List[str],
    rows2: List[Row],
    mapping: List[Tuple[str, str]],
    key1: Optional[str],
    key2: Optional[str],
    normalize: bool,
    name1: str = "file1",
    name2: str = "file2",
) -> Report:
    """
    Returns a list of dicts, one per compared cell:
      row | key | col_file1 | col_file2 | value_file1 | value_file2 | match | note
    """
    cols1 = [c for c, _ in mapping]
    cols2 = [c for _, c in mapping]

    _check_cols(headers1, cols1 + ([key1] if key1 else []), name1)
    _check_cols(headers2, cols2 + ([key2] if key2 else []), name2)

    only_in_1 = f"only in {name1}"
    only_in_2 = f"only in {name2}"

    results: Report = []

    if key1 and key2:
        # Build lookup: key_value → row for file2
        lookup: Dict[str, Row] = {}
        for r in rows2:
            k = r.get(key2, "")
            lookup[k] = r
        seen_keys = set()

        for i, r1 in enumerate(rows1):
            kval = r1.get(key1, "")
            seen_keys.add(kval)
            r2 = lookup.get(kval)
            for c1, c2 in mapping:
                v1 = r1.get(c1, "")
                if r2 is None:
                    results.append(_cell(i, kval, c1, c2, v1, "", False, only_in_1))
                else:
                    v2 = r2.get(c2, "")
                    match = (_norm(v1) == _norm(v2)) if normalize else (v1 == v2)
                    results.append(_cell(i, kval, c1, c2, v1, v2, match, ""))

        # Rows present only in file2
        for r2 in rows2:
            kval = r2.get(key2, "")
            if kval not in seen_keys:
                for c1, c2 in mapping:
                    v2 = r2.get(c2, "")
                    results.append(_cell("—", kval, c1, c2, "", v2, False, only_in_2))
    else:
        # Positional: compare row i of file1 with row i of file2
        max_len = max(len(rows1), len(rows2))
        for i in range(max_len):
            r1 = rows1[i] if i < len(rows1) else None
            r2 = rows2[i] if i < len(rows2) else None
            for c1, c2 in mapping:
                v1 = r1.get(c1, "") if r1 else ""
                v2 = r2.get(c2, "") if r2 else ""
                if r1 is None:
                    note = only_in_2
                    match = False
                elif r2 is None:
                    note = only_in_1
                    match = False
                else:
                    note = ""
                    match = (_norm(v1) == _norm(v2)) if normalize else (v1 == v2)
                results.append(_cell(i + 1, "", c1, c2, v1, v2, match, note))

    return results


def _cell(row, key, c1, c2, v1, v2, match, note) -> dict:
    return {
        "row": row,
        "key": key,
        "col_file1": c1,
        "col_file2": c2,
        "value_file1": v1,
        "value_file2": v2,
        "match": match,
        "note": note,
    }


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def _print_summary(
    report: Report,
    file1: str,
    file2: str,
    rows1_before: int,
    rows1_after: int,
    rows2_before: int,
    rows2_after: int,
) -> None:
    total      = len(report)
    matched    = sum(1 for r in report if r["match"])
    mismatched = total - matched
    pct        = 100 * matched / total if total else 0

    print()
    print(_c("=" * 62, BOLD))
    print(_c("  Column Comparison Summary", BOLD, CYAN))
    print(_c("=" * 62, BOLD))
    print(f"  File 1 : {_c(file1, BOLD)}")
    if rows1_before != rows1_after:
        print(f"           {_c(f'{rows1_before} rows → {rows1_after} rows after filter', YELLOW)}")
    else:
        print(f"           {rows1_before} rows")
    print(f"  File 2 : {_c(file2, BOLD)}")
    if rows2_before != rows2_after:
        print(f"           {_c(f'{rows2_before} rows → {rows2_after} rows after filter', YELLOW)}")
    else:
        print(f"           {rows2_before} rows")
    print()
    print(f"  Total comparisons : {_c(total, BOLD)}")
    print(f"  Matched           : {_c(matched, BOLD, GREEN)}")
    print(f"  Mismatched        : {_c(mismatched, BOLD, RED)}")
    print(f"  Match rate        : {_c(f'{pct:.1f}%', BOLD)}")
    print(_c("=" * 62, BOLD))


def _print_per_column(report: Report) -> None:
    # Group by (col_file1, col_file2)
    groups: Dict[Tuple, List] = {}
    for r in report:
        key = (r["col_file1"], r["col_file2"])
        groups.setdefault(key, []).append(r)

    print()
    print(_c("  Per-column breakdown", BOLD))
    print(_c("  " + "─" * 58, DIM))
    for (c1, c2), rows in groups.items():
        total   = len(rows)
        matched = sum(1 for r in rows if r["match"])
        pct     = 100 * matched / total if total else 0
        filled  = int(pct / 5)
        bar     = "█" * filled + "░" * (20 - filled)
        colour  = GREEN if pct == 100 else (YELLOW if pct >= 80 else RED)
        label   = f"{c1} → {c2}"
        print(f"  {label:<30}  {_c(bar, colour)}  {matched}/{total} ({pct:.0f}%)")
    print()


def _print_detail(report: Report, mismatches_only: bool) -> None:
    rows = [r for r in report if not r["match"]] if mismatches_only else report
    if not rows:
        print(_c("\n  All compared values match!", GREEN, BOLD))
        return

    header = (
        f"{'ROW':>5}  {'KEY':<18}  {'COL (f1→f2)':<28}"
        f"  {'VALUE FILE1':<24}  {'VALUE FILE2':<24}  STATUS"
    )
    print(_c(header, DIM))
    print(_c("─" * len(header), DIM))

    for r in rows:
        col_pair = f"{r['col_file1']} → {r['col_file2']}"
        status   = _c("✓ match", GREEN) if r["match"] else _c("✗ mismatch", RED)
        note     = f"  [{r['note']}]" if r["note"] else ""
        v1       = str(r["value_file1"])[:23]
        v2       = str(r["value_file2"])[:23]
        key      = str(r["key"])[:17]
        print(
            f"{str(r['row']):>5}  {key:<18}  {col_pair:<28}"
            f"  {v1:<24}  {v2:<24}  {status}{note}"
        )


# ---------------------------------------------------------------------------
# Output — save report to CSV or XLSX
# ---------------------------------------------------------------------------

REPORT_FIELDS = ["row", "key", "col_file1", "col_file2",
                 "value_file1", "value_file2", "match", "note"]


def _save_csv(report: Report, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=REPORT_FIELDS)
        w.writeheader()
        w.writerows(report)


def _save_xlsx(report: Report, path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.exit(
            "openpyxl is required to write XLSX files.\n"
            "Install it with:  pip install openpyxl"
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    header_font  = Font(bold=True)
    match_fill   = PatternFill("solid", fgColor="C6EFCE")
    mismatch_fill = PatternFill("solid", fgColor="FFC7CE")

    ws.append(REPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font

    for r in report:
        ws.append([r[f] for f in REPORT_FIELDS])
        fill = match_fill if r["match"] else mismatch_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Auto-width
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 40)

    wb.save(path)


def _save_report(report: Report, path: str) -> None:
    ext = Path(path).suffix.lower()
    try:
        if ext == ".csv":
            _save_csv(report, path)
        elif ext in {".xlsx", ".xls"}:
            _save_xlsx(report, path)
        else:
            sys.exit(f"Unsupported output format '{ext}'. Use .csv or .xlsx.")
        print(_c(f"\n  Report saved → {path}", BOLD, CYAN))
    except Exception as exc:
        sys.exit(f"Could not write {path}: {exc}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="compare_columns.py",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("file1", help="First file (CSV or XLSX)")
    p.add_argument("file2", help="Second file (CSV or XLSX)")
    p.add_argument(
        "--map", "-m",
        nargs="+",
        required=True,
        metavar="COL1:COL2",
        help="Column mapping pairs: 'file1_col:file2_col'. Repeat for multiple pairs.",
    )

    kg = p.add_argument_group("row matching (optional — defaults to positional)")
    kg.add_argument("--key",  metavar="COL", help="Key column when it has the same name in both files.")
    kg.add_argument("--key1", metavar="COL", help="Key column in file1.")
    kg.add_argument("--key2", metavar="COL", help="Key column in file2.")

    p.add_argument("--sheet1", metavar="SHEET", help="Sheet name/index for file1 (XLSX only).")
    p.add_argument("--sheet2", metavar="SHEET", help="Sheet name/index for file2 (XLSX only).")

    fg = p.add_argument_group(
        "row filtering (reduce comparison range before comparing)",
        description=(
            "Filter expressions: col=value  col!=value  col~value  "
            "col!~value  col^value  col$value\n"
            "Multiple filters on the same file are ANDed together."
        ),
    )
    fg.add_argument(
        "--filter1",
        nargs="+",
        metavar="EXPR",
        default=[],
        help="Filter rows in file1 before comparing. E.g. --filter1 \"cust_no=40007\"",
    )
    fg.add_argument(
        "--filter2",
        nargs="+",
        metavar="EXPR",
        default=[],
        help="Filter rows in file2 before comparing. E.g. --filter2 \"Section~produce\"",
    )

    p.add_argument("--normalize",       action="store_true", help="Ignore case and surrounding whitespace.")
    p.add_argument("--mismatches-only", action="store_true", help="Only show mismatched rows in the detail table.")
    p.add_argument("--output", "-o",    metavar="FILE", help="Save diff report to a CSV or XLSX file.")
    p.add_argument(
        "--list-columns",
        action="store_true",
        help="Print column names of both files and exit (handy before building --map).",
    )
    return p


def main() -> None:
    args = _build_parser().parse_args()

    key1 = args.key1 or args.key
    key2 = args.key2 or args.key

    headers1, rows1 = _load(args.file1, args.sheet1)
    headers2, rows2 = _load(args.file2, args.sheet2)

    if args.list_columns:
        print(_c(f"\nColumns in {args.file1}:", BOLD, CYAN))
        for i, h in enumerate(headers1, 1):
            print(f"  {i:>3}. {h}")
        print(_c(f"\nColumns in {args.file2}:", BOLD, CYAN))
        for i, h in enumerate(headers2, 1):
            print(f"  {i:>3}. {h}")
        print()
        return

    # Parse and apply filters
    filters1 = _parse_filters(args.filter1, headers1, f"file1 ({args.file1})")
    filters2 = _parse_filters(args.filter2, headers2, f"file2 ({args.file2})")

    rows1_before = len(rows1)
    rows2_before = len(rows2)
    rows1 = _apply_filters(rows1, filters1)
    rows2 = _apply_filters(rows2, filters2)

    if filters1:
        print(_c(f"  file1 filter: {rows1_before} → {len(rows1)} rows", CYAN))
    if filters2:
        print(_c(f"  file2 filter: {rows2_before} → {len(rows2)} rows", CYAN))

    mapping = _parse_mapping(args.map)
    report  = compare(
        headers1, rows1, headers2, rows2, mapping, key1, key2, args.normalize,
        name1=Path(args.file1).name,
        name2=Path(args.file2).name,
    )

    _print_summary(report, args.file1, args.file2, rows1_before, len(rows1), rows2_before, len(rows2))
    _print_per_column(report)
    _print_detail(report, args.mismatches_only)

    if args.output:
        _save_report(report, args.output)

    print()


if __name__ == "__main__":
    main()
